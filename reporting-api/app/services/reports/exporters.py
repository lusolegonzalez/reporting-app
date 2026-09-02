"""Exportadores de ReportResponse a Excel y PDF.

Módulo de bajo acoplamiento: recibe un ReportResponse ya resuelto y
devuelve bytes listos para enviar como respuesta HTTP descargable.

Librerías:
- Excel: openpyxl (puro Python, sin dependencias nativas).
- PDF:   reportlab (ampliamente usado en el ecosistema Python).
"""
from __future__ import annotations

import io
from datetime import date, datetime
from pathlib import Path
from typing import Any

from app.services.reports.base import ReportResponse, ReportSection

# Ruta al logo para incluir en el PDF (ubicado en app/resources/)
_LOGO_PATH = Path(__file__).parent.parent.parent / "resources" / "logo-south-agribusiness.png"
_IFSSA_LOGO_PATH = Path(__file__).parent.parent.parent / "resources" / "ifssa.png"


# ---------------------------------------------------------------------------
# Helpers comunes
# ---------------------------------------------------------------------------


def _make_filename(response: ReportResponse, ext: str) -> str:
    """Nombre de archivo con el código del reporte y fechas si están disponibles."""
    base = response.codigo_reporte.lower()
    desde = response.parametros.get("fecha_desde")
    hasta = response.parametros.get("fecha_hasta")
    if desde and hasta:
        suffix = f"_{desde}_{hasta}"
    else:
        suffix = f"_{response.generado_en.strftime('%Y-%m-%d')}"
    return f"{base}{suffix}.{ext}"


def _fmt_param(val: Any) -> str:
    if isinstance(val, bool):
        return "Sí" if val else "No"
    if isinstance(val, (date, datetime)):
        return val.isoformat()
    return str(val)


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------


def export_to_excel(response: ReportResponse) -> bytes:
    """Genera un .xlsx a partir de un ReportResponse ya resuelto."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)  # quitar hoja en blanco por defecto

    # ── Hoja Resumen ────────────────────────────────────────────────────────
    ws = wb.create_sheet("Resumen")

    row = 1
    cell = ws.cell(row=row, column=1, value=response.nombre_reporte)
    cell.font = Font(bold=True, size=14)
    row += 1
    ws.cell(
        row=row,
        column=1,
        value=f"Generado: {response.generado_en.strftime('%d/%m/%Y %H:%M')}",
    ).font = Font(italic=True, color="888888")
    row += 2

    # Parámetros
    ws.cell(row=row, column=1, value="Parámetros").font = Font(bold=True, size=11)
    row += 1
    for key, val in response.parametros.items():
        ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row, column=2, value=_fmt_param(val))
        row += 1
    row += 1

    # Alertas: se omiten deliberadamente en la exportacion (Excel/PDF).
    # Las alertas solo se muestran en pantalla; ver ReportRunner.tsx.

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55

    # ── Una hoja por sección ─────────────────────────────────────────────────
    for seccion in response.secciones:
        _excel_section_sheet(wb, seccion)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _excel_section_sheet(wb: Any, seccion: ReportSection) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    # Nombre de hoja: max 31 chars, sin chars prohibidos
    _INVALID = r"\/*?:[]"
    sheet_name = seccion.titulo
    for ch in _INVALID:
        sheet_name = sheet_name.replace(ch, "-")
    sheet_name = sheet_name[:31]
    ws = wb.create_sheet(sheet_name)

    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    TOTAL_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    TOTAL_FONT = Font(bold=True)
    LABEL_FONT = Font(bold=True)

    num_cols = len(seccion.columnas)

    row = 1
    # Título de la sección
    ws.cell(row=row, column=1, value=seccion.titulo).font = Font(bold=True, size=12)
    if num_cols > 1:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    row += 2

    # Encabezados de columna
    for col_idx, col in enumerate(seccion.columnas, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col["titulo"])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Filas de datos
    for fila in seccion.filas:
        for col_idx, col in enumerate(seccion.columnas, start=1):
            val = fila.get(col["key"])
            cell = ws.cell(row=row, column=col_idx, value=val)
            if col.get("tipo") == "number" and isinstance(val, (int, float)):
                cell.number_format = "#,##0.000"
                cell.alignment = Alignment(horizontal="right")
        row += 1

    if not seccion.filas:
        ws.cell(row=row, column=1, value="Sin datos para los parámetros indicados.").font = Font(
            italic=True, color="888888"
        )
        row += 1

    # Fila de totales (cajas, kg_neto y demás claves que coincidan con columnas)
    totales = seccion.totales or {}
    col_keys = [col["key"] for col in seccion.columnas]
    totales_en_columnas = {k: v for k, v in totales.items() if k in col_keys and k != "tropas"}
    totales_extra = {
        k: v for k, v in totales.items() if k not in col_keys and k != "tropas"
    }

    if totales_en_columnas or totales_extra:
        row += 1
        ws.cell(row=row, column=1, value="TOTALES").font = TOTAL_FONT
        for col_idx, col in enumerate(seccion.columnas, start=1):
            val = totales_en_columnas.get(col["key"])
            if val is not None:
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.font = TOTAL_FONT
                cell.fill = TOTAL_FILL
                if isinstance(val, (int, float)):
                    cell.number_format = "#,##0.000"
                    cell.alignment = Alignment(horizontal="right")
        row += 1

        # Totales no-columna (ej. cabezas_faenadas)
        for key, val in totales_extra.items():
            label = key.replace("_", " ").capitalize()
            ws.cell(row=row, column=1, value=f"{label}:").font = LABEL_FONT
            cell = ws.cell(row=row, column=2, value=val)
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0"
            row += 1

    # Tropas
    tropas = totales.get("tropas")
    if isinstance(tropas, list) and tropas:
        row += 1
        ws.cell(row=row, column=1, value="Tropas del día").font = Font(bold=True, size=11)
        row += 1
        ws.cell(row=row, column=1, value="Nro. Tropa").font = LABEL_FONT
        ws.cell(row=row, column=2, value="Cabezas").font = LABEL_FONT
        row += 1
        for tropa in tropas:
            ws.cell(row=row, column=1, value=tropa.get("numero_tropa"))
            ws.cell(row=row, column=2, value=tropa.get("cabezas"))
            row += 1

    # Ancho de columnas automático
    col_widths = [18] * num_cols
    col_widths[0] = 16  # código
    if len(col_widths) > 1:
        col_widths[1] = 35  # descripción
    for idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    # Columna extra para totales no-columna
    from openpyxl.utils import get_column_letter as gcl
    ws.column_dimensions[gcl(num_cols + 1)].width = 22


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------


def export_to_pdf(response: ReportResponse) -> bytes:
    """Genera un .pdf a partir de un ReportResponse ya resuelto."""
    if response.codigo_reporte == "DDJJ_MENUDENCIAS":
        return _export_pdf_ddjj_menudencias(response)
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        HRFlowable,
        Image,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )
    from reportlab.lib.enums import TA_LEFT, TA_CENTER

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=1.8 * cm,
        leftMargin=1.8 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        title=response.nombre_reporte,
    )

    styles = getSampleStyleSheet()
    style_title = ParagraphStyle(
        "ReportTitle",
        parent=styles["Title"],
        fontSize=16,
        spaceAfter=4,
        textColor=colors.HexColor("#1F4E79"),
    )
    style_h2 = ParagraphStyle(
        "ReportH2",
        parent=styles["Heading2"],
        fontSize=12,
        spaceAfter=4,
        spaceBefore=8,
        textColor=colors.HexColor("#1F4E79"),
    )
    style_normal = styles["Normal"]
    style_small = ParagraphStyle("Small", parent=style_normal, fontSize=8, textColor=colors.grey)
    _ALERTA_COLORS = {
        "error": colors.HexColor("#CC0000"),
        "warning": colors.HexColor("#CC6600"),
        "info": colors.HexColor("#0055AA"),
    }

    elements: list[Any] = []

    # ── Encabezado con logo ───────────────────────────────────────────────────
    # Ancho útil en landscape A4: 29.7cm - 2*1.8cm = 26.1cm
    usable_w = 26.1 * cm
    logo_w = 4.5 * cm   # ancho del logo (no muy grande)

    title_block = [
        Paragraph(response.nombre_reporte, style_title),
        Paragraph(
            f"Generado: {response.generado_en.strftime('%d/%m/%Y %H:%M')}",
            style_small,
        ),
    ]

    if _LOGO_PATH.exists():
        logo = Image(str(_LOGO_PATH), width=logo_w, height=logo_w * 0.38)
        header_table = Table(
            [[title_block, logo]],
            colWidths=[usable_w - logo_w - 0.5 * cm, logo_w],
        )
        header_table.setStyle(
            TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                ("TOPPADDING", (0, 0), (-1, -1), 0),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ])
        )
        elements.append(header_table)
    else:
        # Fallback sin logo
        elements.append(Paragraph(response.nombre_reporte, style_title))
        elements.append(Paragraph(
            f"Generado: {response.generado_en.strftime('%d/%m/%Y %H:%M')}",
            style_small,
        ))

    elements.append(Spacer(1, 0.3 * cm))

    # ── Parámetros ───────────────────────────────────────────────────────────
    elements.append(Paragraph("Parámetros", style_h2))
    param_data = [
        [
            Paragraph("<b>Parámetro</b>", style_normal),
            Paragraph("<b>Valor</b>", style_normal),
        ]
    ]
    for key, val in response.parametros.items():
        param_data.append([key, _fmt_param(val)])

    param_table = Table(param_data, colWidths=[5 * cm, 8 * cm], hAlign="LEFT")
    param_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2F8")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    elements.append(param_table)

    # Alertas: se omiten deliberadamente en la exportacion (Excel/PDF).
    # Solo se muestran en pantalla (ReportRunner.tsx).

    # ── Secciones ─────────────────────────────────────────────────────────────
    for seccion in response.secciones:
        elements.extend(_pdf_section(seccion, style_h2, style_normal, style_small))

    doc.build(elements)
    return buf.getvalue()


def _pdf_section(
    seccion: ReportSection,
    style_h2: Any,
    style_normal: Any,
    style_small: Any,
) -> list[Any]:
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import HRFlowable, Paragraph, Spacer, Table, TableStyle

    elements: list[Any] = []
    elements.append(Spacer(1, 0.4 * cm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.HexColor("#CCCCCC")))
    elements.append(Paragraph(seccion.titulo, style_h2))

    # Datos de la sección
    if not seccion.filas:
        elements.append(
            Paragraph("Sin datos para los parámetros indicados.", style_small)
        )
    else:
        # Construir tabla: headers + filas + fila de totales
        headers = [col["titulo"] for col in seccion.columnas]
        col_types = [col.get("tipo", "string") for col in seccion.columnas]
        col_keys = [col["key"] for col in seccion.columnas]

        # Calcular anchos de columna (landscape A4 usable ~24 cm)
        usable_width = 24 * cm
        n = len(headers)
        # Primera col (código) estrecha, segunda (desc) ancha, resto iguales
        if n == 4:
            col_widths = [3.5 * cm, 9 * cm, 3.5 * cm, 4.5 * cm]
        elif n == 2:
            col_widths = [8 * cm, 8 * cm]
        else:
            col_widths = [usable_width / n] * n

        def _cell_style(val: Any, tipo: str) -> str:
            if tipo == "number":
                if val is None:
                    return "-"
                try:
                    return f"{float(val):,.3f}".replace(",", "X").replace(".", ",").replace("X", ".")
                except (TypeError, ValueError):
                    return str(val)
            if val is None:
                return "-"
            return str(val)

        table_data = [headers]
        for fila in seccion.filas:
            table_data.append(
                [_cell_style(fila.get(k), t) for k, t in zip(col_keys, col_types)]
            )

        # Fila de totales
        totales = seccion.totales or {}
        totales_en_cols = {k: v for k, v in totales.items() if k in col_keys and k != "tropas"}
        if totales_en_cols:
            total_row = [""] * n
            total_row[0] = "TOTALES"
            for col_idx, key in enumerate(col_keys):
                if key in totales_en_cols:
                    total_row[col_idx] = _cell_style(totales_en_cols[key], col_types[col_idx])
            table_data.append(total_row)

        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        n_rows = len(table_data)
        n_total_row = n_rows - 1 if totales_en_cols else None

        style_cmds = [
            # Header
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2 if n_total_row else -1), [colors.white, colors.HexColor("#EEF2F8")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            # Align numbers to right (last columns)
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ]

        if n_total_row is not None:
            style_cmds += [
                ("BACKGROUND", (0, n_total_row), (-1, n_total_row), colors.HexColor("#D9E1F2")),
                ("FONTNAME", (0, n_total_row), (-1, n_total_row), "Helvetica-Bold"),
            ]

        table.setStyle(TableStyle(style_cmds))
        elements.append(table)

        # Totales extra (no-columna, ej. cabezas_faenadas)
        totales_extra = {
            k: v for k, v in totales.items() if k not in col_keys and k != "tropas"
        }
        for key, val in totales_extra.items():
            label = key.replace("_", " ").capitalize()
            elements.append(
                Paragraph(f"<b>{label}:</b> {val}", style_small)
            )

    # Tropas
    tropas = (seccion.totales or {}).get("tropas")
    if isinstance(tropas, list) and tropas:
        from reportlab.lib.units import cm
        from reportlab.platypus import Spacer, Table, TableStyle

        elements.append(Spacer(1, 0.3 * cm))
        elements.append(Paragraph("Tropas del día", style_h2))
        tropa_data = [["Nro. Tropa", "Cabezas"]]
        for t in tropas:
            tropa_data.append([str(t.get("numero_tropa", "-")), str(t.get("cabezas", "-"))])
        tropa_table = Table(tropa_data, colWidths=[5 * cm, 4 * cm], hAlign="LEFT")
        tropa_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2F8")]),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        elements.append(tropa_table)

    return elements


# ---------------------------------------------------------------------------
# PDF especializado: DDJJ_MENUDENCIAS (retrato A4, formato SENASA/IFSSA)
# ---------------------------------------------------------------------------


def _export_pdf_ddjj_menudencias(response: ReportResponse) -> bytes:
    """Layout DDJJ_MENUDENCIAS: retrato A4, encabezado IFSSA, sin títulos de bloque."""
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.platypus import (
        HRFlowable,
        Image,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    # ── Footer canvas: "Página X de Y" + área de firma ───────────────────────
    class _NumberedCanvas(pdf_canvas.Canvas):
        def __init__(self, *args, **kwargs):
            pdf_canvas.Canvas.__init__(self, *args, **kwargs)
            self._saved_page_states: list[dict] = []

        def showPage(self):
            self._saved_page_states.append(dict(self.__dict__))
            self._startPage()

        def save(self):
            num_pages = len(self._saved_page_states)
            for i, state in enumerate(self._saved_page_states, 1):
                self.__dict__.update(state)
                self._draw_footer(i, num_pages)
                pdf_canvas.Canvas.showPage(self)
            pdf_canvas.Canvas.save(self)

        def _draw_footer(self, page_num: int, total_pages: int) -> None:
            pw, _ = self._pagesize
            is_last = page_num == total_pages
            self.saveState()

            # ── Línea separadora (todas las páginas) ──────────────────────────
            self.setStrokeColor(colors.HexColor("#1F4E79"))
            self.setLineWidth(0.4)
            self.line(1.5 * cm, 1.6 * cm, pw - 1.5 * cm, 1.6 * cm)

            # ── Firma (última página únicamente) ──────────────────────────────
            # ── Numeración de página (todas las páginas) ──────────────────────
            self.setFont("Helvetica", 7)
            self.setFillColor(colors.HexColor("#555555"))
            if is_last:
                self.drawString(1.5 * cm, 0.5 * cm, "Firma y sello aclaratorio Responsable de la empresa")
            self.drawRightString(pw - 1.5 * cm, 0.5 * cm, f"Página {page_num} de {total_pages}")

            self.restoreState()

    buf = io.BytesIO()
    # Retrato A4; bottomMargin ampliado para el footer
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=2.0 * cm,
        title="DECLARACIÓN JURADA PRODUCCIÓN OBTENIDA MENUDENCIAS",
    )
    usable_w = 18 * cm  # 21cm - 2*1.5cm

    styles = getSampleStyleSheet()
    style_normal = styles["Normal"]

    style_doc_title = ParagraphStyle(
        "DDJJTitle",
        parent=style_normal,
        fontSize=10,
        fontName="Helvetica-Bold",
        alignment=TA_CENTER,
        textColor=colors.HexColor("#1F4E79"),
        leading=13,
    )
    style_meta = ParagraphStyle(
        "DDJJMeta",
        parent=style_normal,
        fontSize=7,
        alignment=TA_RIGHT,
        leading=10,
    )
    style_body = ParagraphStyle(
        "DDJJBody",
        parent=style_normal,
        fontSize=8,
        leading=11,
    )
    style_date_header = ParagraphStyle(
        "DDJJDateHeader",
        parent=style_normal,
        fontSize=8,
        fontName="Helvetica-Bold",
        textColor=colors.white,
        leading=11,
    )
    style_small = ParagraphStyle(
        "DDJJSmall",
        parent=style_normal,
        fontSize=7,
        textColor=colors.grey,
    )
    style_section_lbl = ParagraphStyle(
        "DDJJSectionLbl",
        parent=style_normal,
        fontSize=8,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#1F4E79"),
        spaceBefore=4,
        spaceAfter=1,
    )

    elements: list[Any] = []

    # ── Encabezado: logo | título centrado | metadatos derecha ───────────────
    logo_w = 3.0 * cm
    logo_path = _IFSSA_LOGO_PATH if _IFSSA_LOGO_PATH.exists() else _LOGO_PATH
    if logo_path.exists():
        img_obj = Image(str(logo_path))
        ratio = img_obj.imageHeight / img_obj.imageWidth if img_obj.imageWidth else 0.75
        logo_cell: Any = Image(str(logo_path), width=logo_w, height=logo_w * ratio)
    else:
        logo_cell = Paragraph("IFSSA", style_body)

    emision = response.generado_en.strftime("%d/%m/%Y %H:%M")
    meta_cell = Paragraph(
        f"Código: IFFSA FC PRO<br/>Revisión: 00<br/>Emisión: {emision}",
        style_meta,
    )
    title_cell = Paragraph(
        "DECLARACIÓN JURADA PRODUCCIÓN OBTENIDA MENUDENCIAS",
        style_doc_title,
    )

    meta_w = 4.5 * cm
    title_w = usable_w - logo_w - meta_w
    header_tbl = Table(
        [[logo_cell, title_cell, meta_cell]],
        colWidths=[logo_w, title_w, meta_w],
    )
    header_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (2, 0), (2, 0), "RIGHT"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    elements.append(header_tbl)
    elements.append(Spacer(1, 0.15 * cm))
    elements.append(HRFlowable(width="100%", thickness=1.0, color=colors.HexColor("#1F4E79")))
    elements.append(Spacer(1, 0.2 * cm))

    # ── Tabla de fechas ───────────────────────────────────────────────────────
    p = response.parametros

    def _fd(val: Any) -> str:
        if isinstance(val, (date, datetime)):
            return val.strftime("%d/%m/%Y")
        return str(val) if val else "-"

    date_data = [
        [
            Paragraph("Fecha de", style_date_header),
            Paragraph("Desde", style_date_header),
            Paragraph("Hasta", style_date_header),
        ],
        ["Faena", _fd(p.get("fecha_faena_desde")), _fd(p.get("fecha_faena_hasta"))],
        ["Producción", _fd(p.get("fecha_produccion_desde")), _fd(p.get("fecha_produccion_hasta"))],
    ]
    date_tbl = Table(date_data, colWidths=[3.5 * cm, 3.0 * cm, 3.0 * cm], hAlign="LEFT")
    date_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EEF2F8")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
    ]))
    elements.append(date_tbl)
    elements.append(Spacer(1, 0.2 * cm))

    # ── Cabezas faenadas y tropas ─────────────────────────────────────────────
    cabezas_val = None
    tropas_list = None
    for sec in response.secciones:
        tot = sec.totales or {}
        if cabezas_val is None and "cabezas_faenadas" in tot:
            cabezas_val = tot["cabezas_faenadas"]
        if tropas_list is None and isinstance(tot.get("tropas"), list) and tot["tropas"]:
            tropas_list = tot["tropas"]

    if cabezas_val is not None:
        try:
            f = float(cabezas_val)
            cabezas_str = str(int(round(f)))
        except (TypeError, ValueError):
            cabezas_str = str(cabezas_val)
        elements.append(Paragraph(f"<b>Cabezas Faenadas:</b> {cabezas_str}", style_body))

    if tropas_list:
        numeros = ", ".join(str(t.get("numero_tropa", "")) for t in tropas_list)
        elements.append(Paragraph(f"<b>Tropas:</b> {numeros}", style_body))

    elements.append(Spacer(1, 0.2 * cm))

    # ── Tablas de datos (sin títulos de bloque) ───────────────────────────────
    _SECCION_LABELS = {
        "diaria": "Menudencias",
        "decomisos": "Decomisos",
        "mensual": "Menudencias",
    }
    for sec in response.secciones:
        label = _SECCION_LABELS.get(sec.codigo, sec.codigo.capitalize())
        elements.append(Paragraph(label, style_section_lbl))
        elements.extend(_pdf_section_ddjj(sec, style_body, style_small, usable_w))

    doc.build(elements, canvasmaker=_NumberedCanvas)
    return buf.getvalue()


def _pdf_section_ddjj(
    seccion: ReportSection,
    style_body: Any,
    style_small: Any,
    usable_w: float,
) -> list[Any]:
    """Renderiza una sección como tabla compacta sin título ni tropas (ya mostradas arriba)."""
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle

    elements: list[Any] = []

    if not seccion.filas:
        elements.append(Paragraph("Sin datos para los parámetros indicados.", style_small))
        elements.append(Spacer(1, 0.2 * cm))
        return elements

    col_keys = [c["key"] for c in seccion.columnas]
    col_types = [c.get("tipo", "string") for c in seccion.columnas]
    headers = [c["titulo"] for c in seccion.columnas]
    n = len(headers)

    # Anchos para retrato A4 (usable ~18cm)
    if n == 4:
        col_widths = [3.0 * cm, 8.0 * cm, 3.0 * cm, 3.5 * cm]
    elif n == 2:
        col_widths = [8.0 * cm, 8.0 * cm]
    else:
        col_widths = [usable_w / n] * n

    def _fmt(val: Any, tipo: str, col_key: str = "") -> str:
        if tipo == "number":
            if val is None:
                return "-"
            try:
                f_val = float(val)
                if col_key == "cajas":
                    return str(int(round(f_val)))
                return f"{f_val:,.3f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except (TypeError, ValueError):
                return str(val)
        return "-" if val is None else str(val)

    table_data = [headers]
    for fila in seccion.filas:
        table_data.append([_fmt(fila.get(k), t, k) for k, t in zip(col_keys, col_types)])

    totales = seccion.totales or {}
    totales_en_cols = {k: v for k, v in totales.items() if k in col_keys and k != "tropas"}
    has_total_row = bool(totales_en_cols)
    if has_total_row:
        total_row = [""] * n
        total_row[0] = "TOTALES"
        for i, key in enumerate(col_keys):
            if key in totales_en_cols:
                total_row[i] = _fmt(totales_en_cols[key], col_types[i], key)
        table_data.append(total_row)

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
    n_rows = len(table_data)
    n_total = n_rows - 1 if has_total_row else None

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2 if n_total else -1), [colors.white, colors.HexColor("#EEF2F8")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
    ]
    if n_total is not None:
        style_cmds += [
            ("BACKGROUND", (0, n_total), (-1, n_total), colors.HexColor("#D9E1F2")),
            ("FONTNAME", (0, n_total), (-1, n_total), "Helvetica-Bold"),
        ]
    tbl.setStyle(TableStyle(style_cmds))
    elements.append(tbl)
    elements.append(Spacer(1, 0.2 * cm))
    return elements
